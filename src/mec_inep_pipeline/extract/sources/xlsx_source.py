"""Extração de fontes XLSX (ex.: Sinopses Estatísticas do INEP) como dlt resources.

Usa `openpyxl` em modo `read_only=True`, que faz streaming das linhas da
planilha em vez de carregar o workbook inteiro em memória — importante para
os arquivos de 50-200MB citados nos requisitos do projeto.
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import dlt
from openpyxl import load_workbook

from mec_inep_pipeline.config.settings import REPO_ROOT
from mec_inep_pipeline.load.normalizers import apply_minimal_transformations
from mec_inep_pipeline.utils.logging import get_logger

logger = get_logger(__name__)


def _resolve_local_path(source_config: dict[str, Any]) -> Path:
    """Resolve o caminho local do XLSX a ser lido (mesma lógica do csv_source)."""
    local_path = source_config["location"].get("local_fallback_path")
    if not local_path:
        raise ValueError(
            "source_config['location'] precisa definir 'local_fallback_path' "
            "para extração local (download automático via 'url' ainda não implementado)."
        )
    return REPO_ROOT / local_path


def _read_xlsx_rows(
    path: Path,
    *,
    sheet_name: str | None,
    header_row: int,
    skip_footer_rows: int,
) -> Iterator[dict[str, Any]]:
    """Gera dicts a partir de uma planilha XLSX, usando `header_row` como cabeçalho."""
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        if sheet is None:
            raise ValueError(f"Planilha '{sheet_name}' não encontrada em {path}")

        total_rows = sheet.max_row
        last_data_row = total_rows - skip_footer_rows

        headers: list[str] | None = None
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx < header_row:
                continue
            if row_idx == header_row:
                headers = [
                    str(cell).strip() if cell is not None else f"col_{i}"
                    for i, cell in enumerate(row)
                ]
                continue
            if row_idx > last_data_row:
                break
            if headers is None:
                continue
            yield dict(zip(headers, row, strict=False))
    finally:
        workbook.close()


def _obter_valor_por_indice(row: tuple, col_idx: int | None) -> Any:
    """Retorna o valor da linha para uma coluna 1-based. Se col_idx for None, retorna None."""
    if col_idx is None:
        return None
    idx = col_idx - 1
    if idx < 0 or idx >= len(row):
        return None
    value = row[idx]
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _limpar_id(value: Any) -> Any:
    """Converte IDs float inteiros para int, preservando strings e None."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _construir_mapeamento_unpivot(
    headers: list[str],
    *,
    id_column: str,
    prefix_map: dict[str, str],
) -> dict[str, Any]:
    """
    Monta um mapa a partir dos nomes do cabeçalho.

    Retorna algo como:
    {
        "id_escola": 4,
        "vl_indicador_rend": {2005: 13, 2007: 20, ...},
        "vl_nota_matematica": {2005: 84, ...},
        ...
    }
    """
    mapeamento: dict[str, Any] = {"id_escola": None}
    for col_dest in prefix_map.values():
        mapeamento[col_dest] = {}

    padrao = re.compile(r"^(" + "|".join(re.escape(p) for p in prefix_map) + r")_(\d{4})$")

    for idx_1based, header in enumerate(headers, start=1):
        header_limpo = str(header).strip() if header is not None else ""

        if header_limpo == id_column:
            mapeamento["id_escola"] = idx_1based
            continue

        m = padrao.match(header_limpo)
        if m:
            prefixo = m.group(1)
            ano = int(m.group(2))
            col_dest = prefix_map[prefixo]
            mapeamento[col_dest][ano] = idx_1based

    return mapeamento


def _read_xlsx_rows_unpivot(
    path: Path,
    *,
    sheet_name: str | None,
    header_row: int,
    skip_footer_rows: int,
    id_column: str,
    years: list[int],
    prefix_map: dict[str, str],
) -> Iterator[dict[str, Any]]:
    """
    Lê o XLSX e faz o unpivot linha a linha.

    Para cada escola no arquivo de origem, gera `len(years)` linhas normalizadas,
    uma para cada ano.
    """
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        if sheet is None:
            raise ValueError(f"Planilha '{sheet_name}' não encontrada em {path}")

        total_rows = sheet.max_row
        last_data_row = total_rows - skip_footer_rows

        headers: list[str] | None = None
        mapeamento: dict[str, Any] | None = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx < header_row:
                continue

            if row_idx == header_row:
                headers = [
                    str(cell).strip() if cell is not None else f"col_{i}"
                    for i, cell in enumerate(row)
                ]
                mapeamento = _construir_mapeamento_unpivot(
                    headers,
                    id_column=id_column,
                    prefix_map=prefix_map,
                )
                if mapeamento["id_escola"] is None:
                    raise ValueError(f"Coluna ID '{id_column}' não encontrada no cabeçalho.")
                continue

            if row_idx > last_data_row:
                break

            if mapeamento is None:
                continue

            id_escola = _limpar_id(_obter_valor_por_indice(row, mapeamento["id_escola"]))
            if id_escola is None:
                continue

            for ano in years:
                yield {
                    "id_escola": id_escola,
                    "ano_ref": ano,
                    "vl_indicador_rend": _obter_valor_por_indice(
                        row, mapeamento["vl_indicador_rend"].get(ano)
                    ),
                    "vl_nota_matematica": _obter_valor_por_indice(
                        row, mapeamento["vl_nota_matematica"].get(ano)
                    ),
                    "vl_nota_portugues": _obter_valor_por_indice(
                        row, mapeamento["vl_nota_portugues"].get(ano)
                    ),
                    "vl_nota_media": _obter_valor_por_indice(
                        row, mapeamento["vl_nota_media"].get(ano)
                    ),
                    "vl_observado": _obter_valor_por_indice(
                        row, mapeamento["vl_observado"].get(ano)
                    ),
                    "vl_projecao": _obter_valor_por_indice(row, mapeamento["vl_projecao"].get(ano)),
                }
    finally:
        workbook.close()


def build_xlsx_resource(source_name: str, source_config: dict[str, Any]) -> Any:
    """Constrói um dlt resource para uma fonte do tipo 'xlsx'.

    Args:
        source_name: chave da fonte em config/sources.yaml (usada como nome/tabela).
        source_config: bloco de configuração correspondente em config/sources.yaml.
    """
    fmt = source_config.get("format", {})
    sheet_name = fmt.get("sheet_name")
    header_row = int(fmt.get("header_row", 1))
    skip_footer_rows = int(fmt.get("skip_footer_rows", 0))
    unpivot_cfg = fmt.get("unpivot")
    # Ver comentário equivalente em csv_source.py sobre a anotação explícita Any.
    primary_key: Any = source_config.get("primary_key")
    write_disposition = source_config.get("write_disposition", "replace")
    table_name = source_config.get("destination", {}).get("table", source_name)

    # Prepara os parâmetros do unpivot, se houver
    if unpivot_cfg:
        id_column = str(unpivot_cfg["id_column"])
        years = [int(y) for y in unpivot_cfg["years"]]
        prefix_map = {
            str(prefixo): str(col_dest)
            for prefixo, col_dest in unpivot_cfg["field_prefixes"].items()
        }
        if not prefix_map:
            raise ValueError("Bloco 'unpivot.field_prefixes' não pode estar vazio.")
    else:
        id_column = ""
        years = []
        prefix_map = {}

    @dlt.resource(
        name=source_name,
        table_name=table_name,
        write_disposition=write_disposition,
        primary_key=primary_key,
    )
    def resource() -> Iterator[dict[str, Any]]:
        path = _resolve_local_path(source_config)
        if not path.exists():
            logger.warning(
                "Arquivo XLSX não encontrado em %s — pulando fonte '%s'. "
                "Baixe os dados oficiais do INEP antes de rodar o pipeline.",
                path,
                source_name,
            )
            return

        logger.info(
            "Lendo XLSX de %s (sheet=%r, header_row=%d)",
            path,
            sheet_name,
            header_row,
        )

        if unpivot_cfg:
            logger.info("Usando leitura com unpivot para a fonte '%s'", source_name)
            for row in _read_xlsx_rows_unpivot(
                path,
                sheet_name=sheet_name,
                header_row=header_row,
                skip_footer_rows=skip_footer_rows,
                id_column=id_column,
                years=years,
                prefix_map=prefix_map,
            ):
                yield apply_minimal_transformations(row)
        else:
            for row in _read_xlsx_rows(
                path,
                sheet_name=sheet_name,
                header_row=header_row,
                skip_footer_rows=skip_footer_rows,
            ):
                yield apply_minimal_transformations(row)

    return resource
